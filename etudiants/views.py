from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404
from django.template import loader
from django.urls import reverse
from .models import Etudiant
from groupes.models import Groupe
from dep.models import Departement
from django.shortcuts import render, redirect
from .forms import EtudiantForm
from openpyxl import Workbook
from groupes.models import Groupe
from reportlab.pdfgen import canvas



def etud(request):
    mymembers = Etudiant.objects.select_related('id_groupe').all()
    
    template = loader.get_template('etud.html')
    context = {'mymembers': mymembers, }
    return HttpResponse(template.render(context, request))

def view_etud(request, id_etudiant):
    etudiant = get_object_or_404(Etudiant, pk=id_etudiant)
    context = {'etudiant': etudiant}
    return render(request, 'view_etud.html', context)

def addetud(request):
    if request.method == 'POST':
        form = EtudiantForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('etud')  # Rediriger vers la liste des sujets après l'ajout
    else:
        form = EtudiantForm()
    context = {'form': form}
    return render(request, 'addetud.html', context)


def updateetud(request, id_etudiant):
    etudiant = get_object_or_404(Etudiant, id_etudiant=id_etudiant)

    if request.method == 'POST':
        form = EtudiantForm(request.POST, instance=etudiant)
        if form.is_valid():
            form.save()
            context = {'etudiant': etudiant, 'form': form, 'success': True}
            return render(request, 'updateetud.html', context)
    else:
        form = EtudiantForm(instance=etudiant)
    
    context = {'etudiant': etudiant, 'form': form}
    return render(request, 'updateetud.html', context)

def deleteetud(request, id_etudiant):
  if request.method == 'POST':
    member = Etudiant.objects.get(pk=id_etudiant)
    member.delete()
  return HttpResponseRedirect(reverse('etud'))

def crud_list(request):
    # Récupérez les données du CRUD à exporter
    # data = Etudiant.objects.all()  # Remplacez "VotreModele" par le nom de votre modèle réel
    etudients= Etudiant.objects.select_related('id_groupe','id_departement')
    # Créez un classeur Excel
    workbook = Workbook()

    # Accédez à la feuille active
    sheet = workbook.active

    # Ajoutez des en-têtes de colonnes
    sheet['A1'] = 'Nom'
    sheet['B1'] = 'prenom'
    sheet['C1'] = 'Nom du Group'
    sheet['D1'] = 'Code Departements'
    sheet['E1'] = 'note'
   
    # ...

    # Parcourez les données et ajoutez-les aux lignes suivantes
    row = 2
    for etudient in etudients:
        sheet.cell(row=row, column=1).value = etudient.nom_etudiant
        sheet.cell(row=row, column=2).value = etudient.prenom_etudiant
        sheet.cell(row=row, column=3).value = etudient.id_groupe.nom_groupe
        sheet.cell(row=row, column=4).value = etudient.id_departement.nom_departement
        sheet.cell(row=row, column=5).value = etudient.note
        
        # ...
        row += 1

    # Définissez le nom du fichier de sortie
    filename = 'export_etudinets.xlsx'

    # Répondez avec le fichier Excel en tant que réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={0}'.format(filename)

    # Enregistrez le classeur Excel dans la réponse HTTP
    workbook.save(response)

    return response



def export_pdf(request):
    # Récupérer les données CRUD à exporter
    # ...
    queryset= Etudiant.objects.select_related('id_groupe','id_departement')
    etudients = [etudiant.id_etudiant for etudiant in queryset]

    

    # Créer un objet HttpResponse avec le type de contenu approprié
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="export.pdf"'

    # Créer le canvas PDF
    pdf = canvas.Canvas(response)

    # Générer le contenu PDF à partir des données CRUD
     # Ajoutez des en-têtes de colonnes
    # sheet['A1'] = 'Nom'
    # sheet['B1'] = 'prenom'
    # sheet['C1'] = 'Nom du Group'
    # sheet['D1'] = 'Code Departements'
    # sheet['E1'] = 'note'
   
    # ...

    # Parcourez les données et ajoutez-les aux lignes suivantes
    row = 2
    y=700
    for etudient in etudients:
        pdf.drawString(100, y, (etudient.nom_etudiant))
        # sheet.cell(row=row, column=1).value = etudient.nom_etudiant
        # sheet.cell(row=row, column=2).value = etudient.prenom_etudiant
        # sheet.cell(row=row, column=3).value = etudient.id_groupe.nom_groupe
        # sheet.cell(row=row, column=4).value = etudient.id_departement.nom_departement
        # sheet.cell(row=row, column=5).value = etudient.note
        
        # ...
        row += 1
    # Utilisez les méthodes de ReportLab pour dessiner le contenu du PDF
    # Par exemple :
    pdf.drawString(100, 700, etudients)

    # Enregistrez le canvas PDF et terminez la réponse
    pdf.save()

    return response
    

# Create your views here.
